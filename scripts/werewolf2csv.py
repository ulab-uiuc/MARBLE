import os
import re
import json

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment


def process_shared_memory_in_folder(folder_path: str, xlsx_path: str, model_name: str) -> None:
    """
    1) 在 folder_path 下找 'fullrun' 开头的子文件夹, 进入;
    2) 若该子文件夹下只有一个子目录, 继续进入;
    3) 读取 shared_memory.json, 替换日志中的玩家名字, 在日志中第一次出现 "[continue_game]" 前插分割线;
    4) 扫描 .txt 文件, 读 witch/seer/guard 三类文本;
    5) 把数据写到 xlsx_path 中, 其中:
       - 若 xlsx_path 尚不存在, 新建并写入表头;
       - 若已存在, 读取后追加一行.
    6) 最后用 openpyxl 设置 "event_log" 列的单元格 wrap_text=True, 以便Excel中同一个单元格显示多行.
    """

    # (A) 找一个以 "fullrun" 开头的子目录
    subdirs_full_run = []
    for entry in os.listdir(folder_path):
        fullpath = os.path.join(folder_path, entry)
        if os.path.isdir(fullpath) and entry.startswith("fullrun"):
            subdirs_full_run.append(entry)

    if len(subdirs_full_run) == 0:
        raise RuntimeError(f"在 '{folder_path}' 中找不到以 'fullrun' 开头的子文件夹.")
    elif len(subdirs_full_run) > 1:
        raise RuntimeError(f"在 '{folder_path}' 中找到多个以 'fullrun' 开头的子文件夹: {subdirs_full_run}")
    else:
        folder_path = os.path.join(folder_path, subdirs_full_run[0])

    # (B) 若此文件夹下还有唯一一个子文件夹, 再进一层
    subdirs = [d for d in os.listdir(folder_path)
               if os.path.isdir(os.path.join(folder_path, d))]
    if len(subdirs) == 1:
        folder_path = os.path.join(folder_path, subdirs[0])

    # (1) 读取 shared_memory.json
    json_file = os.path.join(folder_path, "shared_memory.json")
    if not os.path.isfile(json_file):
        pass

    with open(json_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    # players_data -> role_map
    players_data = data["private_state"]["players"]
    role_map = {}
    for player_name, info in players_data.items():
        role_map[player_name] = info["role"]

    # 日志
    private_event_log = data["private_event_log"]
    print("DEBUG/StepA: len(private_event_log) =", len(private_event_log))
    print("DEBUG/StepA tail =", private_event_log[-300:])
    # (2) 替换玩家名
    pattern = r"\b(" + "|".join(role_map.keys()) + r")\b"

    def replace_name_with_role(m):
        name = m.group(0)
        return f"{name}({role_map[name]})"

    replaced_log = re.sub(pattern, replace_name_with_role, private_event_log)

    # (3) 找 "[continue_game]"
    split_token = "[continue_game]"
    if split_token in replaced_log:
        parts = replaced_log.split(split_token, 1)
        prologue = parts[0].rstrip("\n")
        simulation = parts[1].lstrip("\n")
        final_log = (
            prologue
            + "\n\n\n\n\n"
            + "============================================CONTINUE GANE============================================\n\n\n\n\n"
            + "[continue_game]"
            + simulation
        )
    else:
        final_log = replaced_log

    # 额外清理文本中不需要的双引号(若有需要)
    # 并统一换行符
    final_log = final_log.replace('"', '').replace('\r\n', '\n').replace('\r', '\n')
    print("DEBUG/StepB: len(final_log) =", len(final_log))
    print("DEBUG/StepB tail =", final_log[-300:])
    # (4) 扫描 .txt
    witch_text = ""
    seer_text = ""
    guard_text = ""

    for filename in os.listdir(folder_path):
        if not filename.lower().endswith(".txt"):
            continue
        fp = os.path.join(folder_path, filename)
        low = filename.lower()
        with open(fp, "r", encoding="utf-8") as ff:
            content = ff.read()
        # 同样清理
        content = content.replace('"', '').replace('\r\n', '\n').replace('\r', '\n')

        if "witch" in low:
            witch_text = content
        elif "seer" in low:
            seer_text = content
        elif "guard" in low:
            guard_text = content

    # 需要的列
    columns = [
        "folder_path",
        "event_log",
        "witch's thought",
        "seer's thought",
        "guard's thought",
        "model",
        "communication_score",
        "planning_score"
    ]

    communication_score = -1
    planning_score = -1

    # 构造DataFrame, 只有一行
    new_row = {
        "folder_path": folder_path,
        "event_log": final_log,
        "witch's thought": witch_text,
        "seer's thought": seer_text,
        "guard's thought": guard_text,
        "model": model_name,
        "communication_score": communication_score,
        "planning_score": planning_score
    }
    df_new = pd.DataFrame([new_row], columns=columns)
    print("DEBUG/StepC: df_new['event_log'][0] tail =", df_new["event_log"][0][-300:])
    # (5) 写到 xlsx
    if not os.path.exists(xlsx_path):
        # 如果文件不存在, 直接写出(含表头)
        df_new.to_excel(xlsx_path, sheet_name="Sheet1", index=False)
    else:
        # 如果已存在, 先读旧数据, 再append, 再写回
        df_old = pd.read_excel(xlsx_path, sheet_name="Sheet1")
        df_combined = pd.concat([df_old, df_new], ignore_index=True)
        df_combined.to_excel(xlsx_path, sheet_name="Sheet1", index=False)

    # (6) 用 openpyxl 设置 event_log 列 wrap_text=True
    wb = load_workbook(xlsx_path)
    ws = wb["Sheet1"]
    val = ws.cell(row=2, column=2).value  # 假设 event_log 在第2列B列
    print("DEBUG/StepD: readback len =", len(val))
    print("DEBUG/StepD tail =", val[-300:])
    # 找 event_log 列的列号
    col_names = [cell.value for cell in ws[1]]  # 第一行表头
    if "event_log" in col_names:
        col_idx = col_names.index("event_log") + 1  # 列索引(1-based)
        # 设置此列的自动换行
        max_row = ws.max_row
        for row_idx in range(2, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(wrap_text=True)

        # 你也可设置其他列 wrap_text, e.g. if "witch's thought" in col_names, etc.

    # 也可设置列宽, 让多行文本更好看
    # 假设 event_log 是第2列 -> 'B'
    # 这里 col_idx=2 => 'B' ; col_idx=3 => 'C', etc.
    # 你可写个小函数把2 -> 'B'
    if col_idx == 2:
        ws.column_dimensions["B"].width = 80
    # 这里仅示例, 你可针对需要列再写

    wb.save(xlsx_path)

    print(f"处理完成: folder={folder_path}, 模型='{model_name}', 已追加到 {xlsx_path}")


def main(top_folder: str, xlsx_path: str):
    """
    1) 在 top_folder 中，找到 6 个模型子文件夹 (例如 a, b, c, d, e, f)
    2) 每个模型子文件夹内，取前10个模拟结果路径 (sorted)
    3) 以 a1 -> b1 -> c1 -> d1 -> e1 -> f1 -> a2 -> b2 -> ... 的顺序调用 process_shared_memory_in_folder
    """

    # 假设你的 6 个模型子文件夹名是严格的 ["a","b","c","d","e","f"] 或者别的固定名称
    # 如果名字不固定，可根据需要做匹配/排序
    model_names = ["4o","4omini","gpt35","llama31_8b","llama31_70b","llama33"]

    # step0: 收集每个模型子文件夹下的“模拟结果目录”列表
    # --------------------------------------------------
    # 结果将存成: model_to_sims = {
    #    "a": [simPath1, simPath2, ...],   # 最多10个
    #    "b": [...],
    #     ...
    # }
    model_to_sims = {}

    for m in model_names:
        model_dir = os.path.join(top_folder, m)  # 形如 top_folder/a

        if not os.path.isdir(model_dir):
            print(f"[WARN] 模型子文件夹不存在: {model_dir}")
            model_to_sims[m] = []
            continue

        # 读取该模型目录下的所有模拟结果子文件夹
        # 例如 "eval_20250126_135325" 这类
        # 你可以根据实际情况做筛选，比如只要开头是 "eval_"，或 sorting
        # 这里先全部拿来并按名称排序
        all_sims = []
        for entry in os.listdir(model_dir):
            sim_path = os.path.join(model_dir, entry)
            if os.path.isdir(sim_path):
                all_sims.append(sim_path)

        # 排序 (假设字典序满足需求)
        all_sims.sort()

        # 只取前10个
        top10 = all_sims[:10]

        model_to_sims[m] = top10

    # step1: 构造轮次顺序 a1->b1->c1->d1->e1->f1->a2->b2->...
    # ------------------------------------------------------
    # 先把 6个模型的结果都收集到 2维 list:
    #   sims_for[0] = model_to_sims["a"]  (list of up to 10 paths)
    #   sims_for[1] = model_to_sims["b"]
    # ...
    sims_for = [model_to_sims[m] for m in model_names]

    # 这里我们假设6个模型都至少有10条，否则要判空
    # 轮流调:
    #   i=0 => a1 b1 c1 d1 e1 f1
    #   i=1 => a2 b2 c2 d2 e2 f2
    # ...
    # 代码写法:
    max_n = 10  # 只做10轮
    total_sequence = []  # 存 (modelName, simPath)

    for i in range(max_n):
        for j, m in enumerate(model_names):
            sims_list = sims_for[j]
            if i < len(sims_list):
                sim_path = sims_list[i]
                total_sequence.append((m, sim_path))

    # step2: 依次调用 process_shared_memory_in_folder
    # ------------------------------------------------
    # 如果你第一条想要做一些特殊处理(“注意修改第一个log和最后一个模型名”之类的),
    # 可以在循环中做分支:
    for idx, (model_folder, sim_path) in enumerate(total_sequence, start=1):
        # 例子: 如果 idx==1, 修改“第一个log”？
        # 例子: 如果 model_folder==... 做点啥？
        # 这里只是示例:
        print(f"[{idx}/{len(total_sequence)}] calling process for model={model_folder}, sim={sim_path}")
        process_shared_memory_in_folder(sim_path, xlsx_path, model_name=model_folder)

    print("所有模拟结果处理完成！")


if __name__ == "__main__":
    # 假设脚本入口
    # your data set folder, 里面有 a,b,c,d,e,f 六个模型子文件夹
    top_folder = r"werewolf_eval"
    # 结果输出
    xlsx_path = r"output.xlsx"

    main(top_folder, xlsx_path)